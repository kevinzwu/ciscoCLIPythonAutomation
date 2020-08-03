#import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows

# Create new workbook
wb = Workbook()
wb.save('filename')

# Load existing workbook
wb = load_workbook(filename)